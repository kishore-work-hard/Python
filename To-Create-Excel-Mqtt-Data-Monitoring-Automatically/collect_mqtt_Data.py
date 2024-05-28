
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import time
import os
import paho.mqtt.client as mqtt

# Function to get current date in 'YYYY-MM-DD' format
def get_current_date():
    return datetime.now().strftime('%Y-%m-%d')

# Function to get current hour in '1AM', '2AM', ..., '12AM' format
def get_current_hour():
    return datetime.now().strftime('%I%p').lstrip('0')  # Removing leading zero

# Function to create a new sheet with hourly columns
def create_new_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
        sheet = wb[sheet_name]
        for i in range(1, 13):
            col_letter = get_column_letter(i)
            sheet[f'{col_letter}1'] = f'{i}AM'
        for i in range(1, 13):
            col_letter = get_column_letter(i + 12)
            sheet[f'{col_letter}1'] = f'{i}PM'
        wb.save('camera_status.xlsx')

# MQTT message handling
def on_message(client, userdata, msg):
    message = msg.payload.decode()
    current_date = get_current_date()
    current_hour = get_current_hour()

    # Load the workbook
    wb = openpyxl.load_workbook(file_name)

    # Create a new sheet if it doesn't exist
    if current_date not in wb.sheetnames:
        create_new_sheet(wb, current_date)

    # Load the sheet for the current date
    sheet = wb[current_date]

    # Get the column letter for the current hour
    col_num = int(datetime.now().strftime('%I').lstrip('0'))
    if 'PM' in current_hour:
        col_num += 12
    col_letter = get_column_letter(col_num)

    # Write the message to the appropriate cell
    cell = f'{col_letter}2'  # Writing to the second row
    sheet[cell] = message

    # Save the workbook
    wb.save(file_name)

# MQTT setup
broker = "host addtess"
port = 1883
topic = "mqtt/topic/here"

client = mqtt.Client()
client.on_message = on_message

client.connect(broker, port, 60)
client.subscribe(topic)
client.loop_start()

# Check if file exists, otherwise create it
file_name = 'camera_status.xlsx'
if not os.path.exists(file_name):
    wb = Workbook()
    wb.save(file_name)

# Main loop to check the time and update the sheet
while True:
    # Wait until the next hour
    current_time = datetime.now()
    next_hour = (current_time + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
    time_to_wait = (next_hour - current_time).total_seconds()
    time.sleep(time_to_wait)
